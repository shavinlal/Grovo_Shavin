'''
Created on 05-Mar-2018

@author: Sheethu C
'''
import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import xlrd
from BaseTestClass import driver
from urllib import splitvalue


class Delete_Tags_Attributes_Lessons:
    
    def delete_Tag(self):
        
        wait=WebDriverWait(driver, 60)
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/div/section[2]/div/ul/li[1]/a")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/div/section[2]/div/ul/li[1]/a").click()
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/header/div/button")))
        driver.find_element_by_xpath("html/body/div/div/div[3]/div[2]/div/header/div/button").click()
        print "Clicked on Delete Button"
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div/button[1]")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div/button[1]").click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[2]/div/div")))
        time.sleep(3)
        
    def deleteLesson(self):
        
        print "\n\nDeleting all lessons.........."
        wait=WebDriverWait(driver, 80)
        wait.until(EC.visibility_of_element_located((By.XPATH,"//a[@href='/create/lessons']")))
    
        print "Clicking on Lessons button from side menu"
        driver.find_element_by_xpath("//a[@href='/create/lessons']").click()
    
    def deleteItLesson(self):    
        wait=WebDriverWait(driver, 80)
        #wait.until(EC.visibility_of_element_located((By.XPATH,"//tr[1]/td[4]/button[.='Delete']")))
        
        time.sleep(3)
        wait.until(EC.element_to_be_clickable((By.XPATH,"//tr[1]/td[4]/button[.='Delete']")))
        driver.find_element_by_xpath("//tr[1]/td[4]/button[.='Delete']").click()
        delo=wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[2]/button[1]")))
        delo.click()
        wait.until(EC.visibility_of_element_located((By.XPATH,".//*[@id='content']/div/div[2]/div/div/span")))     
    
    
    def deleteAttribute(self):
        
        wait=WebDriverWait(driver, 60)
        driver.find_element_by_xpath("//tr[1]/td[5]/button[.='Delete']").click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div[2]/div/div/div[2]/div[2]/button[1]")))
        wait.until(EC.element_to_be_clickable((By.XPATH,"html/body/div[2]/div/div/div[2]/div[2]/button[1]")))
        driver.find_element_by_xpath("html/body/div[2]/div/div/div[2]/div[2]/button[1]").click()
        
        wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[2]/div/div")))
        time.sleep(3)
     
                
    def mainDeletetag(self):
        try:
            d=Delete_Tags_Attributes_Lessons()
            wait=WebDriverWait(driver, 80)
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a")))
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a").click()
            print "Clicked on admin icon"
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]")))
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]").click()
            print "Clicked on Admin"
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[5]").click()
            print "Clicked on Tag"
            time.sleep(8)
            
            count = driver.find_element_by_xpath("//*[@id='content']/div/div[3]/div[2]/div/div/section[2]/ul/li[1]/span").text
            print count
            
            if count=='0':
                print "No Tags to delete"
            else: 
                wait.until(EC.visibility_of_element_located((By.XPATH,"//li[@class='u-inline-block']/a")))
                ele=driver.find_elements_by_xpath("//li[@class='u-inline-block']/a")
                count=len(ele)
            
                wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/div/section[2]/div/ul/li[1]/a")))
                print "Tag Page Loaded"
                
                for count in range (0,count):
                    d.delete_Tag()
                print "All Tags are deleted"   
        finally:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url) 
            
    def mainDelete_attribute(self):
        try:    
            d=Delete_Tags_Attributes_Lessons()
            wait=WebDriverWait(driver, 60)
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a")))
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/a").click()
            print "Clicked on admin icon"
            wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]")))
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]").click()
            print "Clicked on Admin"
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[1]").click()
            print "Clicked on User Attribute"
            driver.find_element_by_xpath("html/body/div/div/div[3]/div[1]/div/nav/div/div[2]/div[6]/div/ul/li[4]").click()
            print "User attribute Page Loaded"
            time.sleep(8)
            
            count = driver.find_element_by_xpath("//*[@id='content']/div/div[3]/div[2]/div/div/ul/li[1]/span").text
            print count
            
            if count=='0':
                print "No Attributes to delete"
            else:
                wait.until(EC.visibility_of_element_located((By.XPATH,"html/body/div/div/div[3]/div[2]/div/div/div/div[2]/div/table/tbody/tr[1]/td[5]/button")))
                print "Deleting all attributes"
                ele=driver.find_elements_by_xpath("//tbody/tr/td[5]/button")
                count=len(ele)
                
                for count in range (0,count):
                    d.deleteAttribute()
                print "All Attributes are deleted"  
            
        finally:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url) 

            
    def mainDeleteLessons(self):
        try:
            d=Delete_Tags_Attributes_Lessons()
            d.deleteLesson()
            
            wait=WebDriverWait(driver, 60)
            wait.until(EC.visibility_of_element_located((By.XPATH,"//tr[1]/td[4]/button[.='Delete']")))
            ele=driver.find_elements_by_xpath("//tbody/tr/td[4]/button[.='Delete']")
            count=len(ele)
                
            for count in range (0,count):
                d.deleteItLesson()
            print "All Lessons are deleted" 
            
        finally:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url) 
            
    def mainUpdateExcelWithUserDetails(self):
        try:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('CreateuserfromUI')
            
            cell2 = second_sheet.cell(1,4)
            LearnId = cell2.value
            LearnsId = LearnId.split("#")
            emp = LearnsId[0]+"#"
            ids = LearnsId[1]
            empId = int(ids)+1
            LearnerEmpId = emp+str(empId)
            print LearnerEmpId
            
            cell1 = second_sheet.cell(1,3)
            LearnEmailId = cell1.value
            LearnerEmail = LearnEmailId.split("@")
            spitValue = LearnerEmail[0][:4]
            email = spitValue+str(empId)
            id = "@"+LearnerEmail[1]
            LearnerEmailId = email+id
            print LearnerEmailId
            
            
            cell4 = second_sheet.cell(2,4)
            CreatorId = cell4.value
            CreatorsId = CreatorId.split("#")
            emp1 = CreatorsId[0]+"#"
            ids1 = CreatorsId[1]
            empId1 = int(ids1)+1
            CreatorsEmpId = emp1+str(empId1)
            print CreatorsEmpId
            
            
            cell3 = second_sheet.cell(2,3)
            CreateEmailId = cell3.value
            CreatorEmail = CreateEmailId.split("@")
            spitValue1 = CreatorEmail[0][:4]
            email1 = spitValue1+str(empId1)
            id1 = "@"+CreatorEmail[1]
            CreatorEmailId = email1+id1
            print CreatorEmailId
            
            cell6 = second_sheet.cell(3,4)
            MasterAdminId = cell6.value
            MasterAdminsId = MasterAdminId.split("#")
            emp2 = MasterAdminsId[0]+"#"
            ids2 = MasterAdminsId[1]
            empId2 = int(ids2)+1
            MasterAdminsEmpId = emp2+str(empId2)
            print MasterAdminsEmpId
            
            cell5 = second_sheet.cell(3,3)
            MasterEmail = cell5.value
            MasterAdminEmail = MasterEmail.split("@")
            spitValue2 = MasterAdminEmail[0][:4]
            email2 = spitValue2+str(empId2)
            id2 = "@"+MasterAdminEmail[1]
            MasterAdminEmailId = email2+id2
            print MasterAdminEmailId
            
            cell8 = second_sheet.cell(4,4)
            LearningAdminId = cell8.value
            LearningAdminsId = LearningAdminId.split("#")
            emp3 = LearningAdminsId[0]+"#"
            ids3 = LearningAdminsId[1]
            empId3 = int(ids3)+1
            LearningAdminsEmpId = emp3+str(empId3)
            print LearningAdminsEmpId
            
            cell7 = second_sheet.cell(4,3)
            LearningEmail = cell7.value
            LearningAdminEmail = LearningEmail.split("@")
            spitValue3 = LearningAdminEmail[0][:4]
            email3 = spitValue3+str(empId3)
            id3 = "@"+LearningAdminEmail[1]
            LearningAdminEmailId = email3+id3
            print LearningAdminEmailId
            
            wb = load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            print (wb.sheetnames)
        
            sheet = wb['CreateuserfromUI']
            
            sheet.cell(row=2, column=4).value = LearnerEmailId
            sheet.cell(row=2, column=5).value = LearnerEmpId
            
            sheet.cell(row=3, column=4).value = CreatorEmailId
            sheet.cell(row=3, column=5).value = CreatorsEmpId
            
            sheet.cell(row=4, column=4).value = MasterAdminEmailId
            sheet.cell(row=4, column=5).value = MasterAdminsEmpId
            
            sheet.cell(row=5, column=4).value = LearningAdminEmailId
            sheet.cell(row=5, column=5).value = LearningAdminsEmpId
            
            wb.save(os.path.abspath(os.path.join(os.path.dirname(__file__),'TestData.xlsx')))
            print "All User Data Updated in Excel" 
            
        finally:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url) 
     
    def mainDeleteAll(self):
        try:     
            d=Delete_Tags_Attributes_Lessons()
            d.mainDeletetag()
            d.mainDelete_attribute()
            d.mainDeleteLessons()
            d.mainUpdateExcelWithUserDetails()
            
        finally:
            book=xlrd.open_workbook(os.path.join('TestData.xlsx'))
            second_sheet = book.sheet_by_name('Login_Credentials')
            cell = second_sheet.cell(1,1)
            url = cell.value
            driver.get(url)            